"""
Export the live model to a workbook.

    python -m pipeline.to_excel

This is the descendant of the original hand-maintained spreadsheet, with the
manual parts removed. Everything that used to be typed in -- the schedule, the
odds, the power ratings, the final scores -- is now fetched and solved, and this
file is a snapshot of the result plus a live P/L layer.

One thing is deliberately NOT carried over: the old Settings sheet was an input,
where editing a cell changed the model. Here it is a read-only record of the
settings that produced this file, because the model runs in the pipeline, not in
Excel. Editing a number here would look like it did something and wouldn't.
The place to change the model is config/settings.json.
"""

from __future__ import annotations

import datetime as dt
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "site", "data")

INK = "1F2430"
MUTED = "6F7887"
FLAG = "9A6B08"
GOODC = "1D7A4C"
BADC = "B02B32"
RULE = Side(style="thin", color="D6DAE3")

F_TITLE = Font(name="Arial", size=15, bold=True, color=INK)
F_SUB = Font(name="Arial", size=9, italic=True, color=MUTED)
F_HEAD = Font(name="Arial", size=9, bold=True, color="FFFFFF")
F_BODY = Font(name="Arial", size=10, color=INK)
F_MUTED = Font(name="Arial", size=9, color=MUTED)
F_NUM = Font(name="Arial", size=10, color=INK)
F_BOLD = Font(name="Arial", size=10, bold=True, color=INK)
F_KPI = Font(name="Arial", size=16, bold=True, color=INK)
FILL_HEAD = PatternFill("solid", fgColor="2A3340")
FILL_BAND = PatternFill("solid", fgColor="F2F3F6")
FILL_NOTE = PatternFill("solid", fgColor="FFF8E6")


def _read(name, default):
    p = os.path.join(DATA, name + ".json")
    if not os.path.exists(p):
        return default
    with open(p, encoding="utf-8") as fh:
        return json.load(fh)


def _title(ws, title, sub):
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    ws["A2"] = sub
    ws["A2"].font = F_SUB
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 26
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")


def _table(ws, top_row, headers, rows, widths, fmts=None, name=None):
    """Write a header row plus data, styled, and return the last row index."""
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=top_row, column=j, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = Border(bottom=RULE)
    for i, row in enumerate(rows):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=top_row + 1 + i, column=j, value=v)
            c.font = F_NUM if isinstance(v, (int, float)) else F_BODY
            c.border = Border(bottom=RULE)
            if fmts and fmts[j - 1]:
                c.number_format = fmts[j - 1]
            if i % 2 == 1:
                c.fill = FILL_BAND
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    last = top_row + len(rows)
    if name and rows:
        ref = f"A{top_row}:{get_column_letter(len(headers))}{last}"
        t = Table(displayName=name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
        ws.add_table(t)
    ws.freeze_panes = ws.cell(row=top_row + 1, column=1)
    return last


def build(path: str) -> str:
    meta = _read("meta", {})
    summary = _read("summary", {})
    board = _read("board", [])
    ledger = _read("ledger", [])
    ratings = _read("ratings", [])
    games = _read("games", [])
    history = _read("model_history", {})
    cfg = meta.get("settings", {})
    symb = cfg.get("currency_symbol", "$")
    money = f'"{symb}"#,##0.00;("{symb}"#,##0.00);"—"'
    pctf = "0.0%"

    wb = Workbook()

    # ---------------------------------------------------------------- Read Me
    ws = wb.active
    ws.title = "Read Me"
    demo = bool(meta.get("demo"))
    _title(ws, "NCAAF Edge — Model Export",
           (f"SIMULATED DATA — generated {meta.get('generated_at','—')} by tools/make_demo.py to show "
            "the layout. The teams are real; every game, line and result is invented. Run "
            "`python -m pipeline.build --full` and re-export to replace it with live ESPN data."
            if demo else
            f"Generated {meta.get('generated_at','—')} from live ESPN data. "
            "This file is a snapshot, not the model. Re-run the pipeline to refresh it."))
    if demo:
        ws["A2"].font = Font(name="Arial", size=9, bold=True, color=FLAG)
        ws["A2"].fill = FILL_NOTE
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    notes = [
        ("What changed", "The schedule, scores, odds, power ratings and rest days are fetched and "
                         "solved automatically. Nothing in this workbook is typed in by hand."),
        ("Power ratings", "Solved from results by ridge-regularised least squares on margin of victory, "
                          "with strength of schedule handled structurally rather than as an adjustment. "
                          "They update themselves every time a game finishes."),
        ("Probabilities", "Margins are modelled with a discrete distribution bumped at football's key "
                          "numbers (3, 7, 10, 14), not a smooth normal curve, so pushes and the value of "
                          "a half-point off 3 are priced correctly."),
        ("Edge", "Model probability minus the break-even probability of the price on offer. "
                 f"A {cfg.get('model',{}).get('selection_haircut',0)*100:.1f}% haircut is applied first, "
                 "because you only bet where the model disagrees with the market, which is also where "
                 "the model's own error is largest."),
        ("Tiers", f"BEST BET ≥ {cfg.get('tiers',{}).get('best_bet',0)*100:.0f}% edge · "
                  f"GOOD ≥ {cfg.get('tiers',{}).get('good',0)*100:.0f}% · "
                  f"LEAN ≥ {cfg.get('tiers',{}).get('lean',0)*100:.0f}% · PASS below. "
                  "Thresholds are raised when the model has thin data on a matchup."),
        ("Settings sheet", "Read-only. It records the settings that produced this file. Editing a value "
                           "there changes nothing — the model runs in the pipeline, not in Excel. "
                           "Change config/settings.json and re-run."),
        ("Bet Ledger", "The only live part. Stake and Result are editable; Profit/Loss, Running Bankroll "
                       "and the Dashboard recalculate from them, so you can correct a settlement by hand."),
        ("Honest warning", "Model edges reflect the assumptions on the Settings sheet, not certainty. "
                           "Check the calibration table before trusting a number, and bet only what you "
                           "can afford to lose."),
    ]
    r = 4
    for k, v in notes:
        ws.cell(row=r, column=1, value=k).font = F_BOLD
        c = ws.cell(row=r, column=2, value=v)
        c.font = F_BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1

    # -------------------------------------------------------------- Dashboard
    ws = wb.create_sheet("Dashboard")
    _title(ws, "Dashboard", "Every figure below is a formula over the Bet Ledger sheet — "
                            "correct a stake or a result there and these follow.")
    n = max(len(ledger), 1)
    lo, hi = 6, 5 + n          # Bet Ledger data rows
    kpis = [
        ("Starting Bankroll", cfg.get("bankroll", {}).get("starting", 0), money),
        ("Current Bankroll", f"=Settings!B5+SUMIF('Bet Ledger'!$G${lo}:$G${hi},\"<>Pending\",'Bet Ledger'!$H${lo}:$H${hi})", money),
        ("Total Staked", f"=SUMIF('Bet Ledger'!$G${lo}:$G${hi},\"<>Pending\",'Bet Ledger'!$F${lo}:$F${hi})", money),
        ("Profit / Loss", f"=SUMIF('Bet Ledger'!$G${lo}:$G${hi},\"<>Pending\",'Bet Ledger'!$H${lo}:$H${hi})", money),
        ("ROI", "=IFERROR(B7/B6,\"\")", pctf),
        ("Bets Settled", f'=COUNTIFS(\'Bet Ledger\'!$G${lo}:$G${hi},"<>Pending",\'Bet Ledger\'!$G${lo}:$G${hi},"<>")', "0"),
        ("Wins", f'=COUNTIF(\'Bet Ledger\'!$G${lo}:$G${hi},"Win")', "0"),
        ("Losses", f'=COUNTIF(\'Bet Ledger\'!$G${lo}:$G${hi},"Loss")', "0"),
        ("Pushes", f'=COUNTIF(\'Bet Ledger\'!$G${lo}:$G${hi},"Push")', "0"),
        ("Win Rate", "=IFERROR(B10/(B10+B11),\"\")", pctf),
        ("Pending", f'=COUNTIF(\'Bet Ledger\'!$G${lo}:$G${hi},"Pending")', "0"),
        ("Live Plays On The Board", sum(1 for b in board if b.get("tier") != "PASS"), "0"),
        ("Solved Home Field (pts)", meta.get("home_field_advantage"), "0.00"),
        ("Brier Score", meta.get("brier"), "0.0000"),
    ]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    r = 4
    for label, val, fmt in kpis:
        ws.cell(row=r, column=1, value=label).font = F_BOLD
        c = ws.cell(row=r, column=2, value=val)
        c.font = F_KPI if r < 9 else F_NUM
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")
        r += 1

    plays = [b for b in board if b.get("tier") != "PASS"][:15]
    ws.cell(row=r + 1, column=1, value="TOP PLAYS ON THE BOARD RIGHT NOW").font = F_BOLD
    _table(ws, r + 2,
           ["Tier", "Game", "Market", "Pick", "Price", "Edge", "Model", "Stake", "Kickoff"],
           [[b["tier"], b["matchup"], b["market"], b["pick"], b["price"], b["edge"],
             b["model_prob"], None, (b.get("game_date") or "")[:16].replace("T", " ")]
            for b in plays],
           [11, 20, 9, 22, 9, 9, 9, 11, 18],
           [None, None, None, None, "+0;-0", "+0.0%;-0.0%", pctf, money, None])

    # --------------------------------------------------------------- The Board
    ws = wb.create_sheet("Board")
    _title(ws, "Full Board",
           "Every market on every upcoming game, priced against the book. Passes are included on "
           "purpose — a model you only see when it likes something cannot be audited. "
           "'Fair' is the market's own view with the vig removed; 'Break-even' is what the price requires.")
    _table(ws, 4,
           ["Tier", "Game", "Kickoff", "Wk", "Market", "Pick", "Line", "Price", "Model",
            "Fair", "Break-even", "Edge", "EV", "Push", "Conf.", "Book", "Why not"],
           [[b.get("tier"), b.get("matchup"), (b.get("game_date") or "")[:16].replace("T", " "),
             b.get("week"), b.get("market"), b.get("pick"), b.get("line"), b.get("price"),
             b.get("model_prob"), b.get("market_fair_prob"), b.get("breakeven"), b.get("edge"),
             b.get("ev"), b.get("push_prob"), b.get("confidence"), b.get("book"),
             b.get("filtered")] for b in board],
           [11, 18, 17, 5, 9, 22, 8, 8, 9, 9, 11, 9, 9, 8, 8, 14, 38],
           [None, None, None, "0", None, None, "+0.0;-0.0", "+0;-0", pctf, pctf, pctf,
            "+0.0%;-0.0%", "+0.0%;-0.0%", pctf, "0%", None, None],
           name="BoardTable")

    # -------------------------------------------------------------- Bet Ledger
    ws = wb.create_sheet("Bet Ledger")
    _title(ws, "Bet Ledger",
           "Logged at the number the bet qualified at, graded once the game finished, never re-priced "
           "afterwards. Stake (F) and Result (G) are yours to correct; Profit/Loss, Running Bankroll "
           "and Cumulative ROI are formulas that follow them.")
    rows = []
    for b in ledger:
        rows.append([(b.get("game_date") or "")[:10], b.get("matchup"), b.get("market"),
                     b.get("pick"), b.get("price"), b.get("stake"), b.get("result"),
                     None, None, None, b.get("tier"), b.get("edge"), b.get("clv_prob"),
                     b.get("final_score"), b.get("book")])
    last = _table(ws, 5,
                  ["Date", "Game", "Market", "Pick", "Price", "Stake", "Result", "Profit/Loss",
                   "Running Bankroll", "Cum. ROI", "Tier", "Edge", "CLV", "Final Score", "Book"],
                  rows,
                  [11, 18, 9, 22, 8, 10, 9, 12, 16, 9, 11, 9, 9, 18, 14],
                  [None, None, None, None, "+0;-0", money, None, money, money, "+0.0%;-0.0%",
                   None, "+0.0%;-0.0%", "+0.0%;-0.0%", None, None])
    for i in range(len(rows)):
        r = 6 + i
        ws.cell(row=r, column=8,
                value=f'=IF(OR($G{r}="",$G{r}="Pending"),"",ROUND(IF($G{r}="Win",'
                      f'IF($E{r}>0,$F{r}*$E{r}/100,$F{r}*100/-$E{r}),IF($G{r}="Loss",-$F{r},0)),2))'
                ).number_format = money
        ws.cell(row=r, column=9, value=f"=Settings!$B$5+SUM($H$6:$H{r})").number_format = money
        ws.cell(row=r, column=10,
                value=f'=IFERROR(SUM($H$6:$H{r})/SUMIF($G$6:$G{r},"<>Pending",$F$6:$F{r}),"")'
                ).number_format = "+0.0%;-0.0%"
    ws.cell(row=4, column=1,
            value="Profit/Loss uses American-odds payout on the price in column E. "
                  "A Push returns the stake, so it books as zero.").font = F_MUTED

    # ----------------------------------------------------------- Power Ratings
    ws = wb.create_sheet("Power Ratings")
    _title(ws, "Power Ratings",
           "Solved from results, not entered. A rating is points better than an average FBS team on a "
           "neutral field, so the gap between two ratings is a projected margin. Offence and defence are "
           "points above league average, opponent-adjusted, with defence signed so positive is good. "
           "ATS record is context only — the model does not bet on it.")
    _table(ws, 4, ["#", "Team", "Rating", "Offence", "Defence", "GP", "ATS L5"],
           [[i + 1, r_["team"], r_.get("rating"), r_.get("off"), r_.get("def"), r_.get("games"),
             (f"{r_['ats']['w']}-{r_['ats']['l']}" + (f"-{r_['ats']['p']}" if r_["ats"].get("p") else ""))
             if r_.get("ats") else "—"] for i, r_ in enumerate(ratings)],
           [6, 10, 10, 10, 10, 6, 10],
           ["0", None, "+0.0;-0.0", "+0.0;-0.0", "+0.0;-0.0", "0", None],
           name="RatingsTable")

    # ------------------------------------------------------- Games and Results
    ws = wb.create_sheet("Games & Results")
    _title(ws, "Games & Results",
           "Every FBS game pulled this season, in order, with the line as last seen — future weeks "
           "included. A blank odds cell means no book has posted that line yet; it is left blank on "
           "purpose rather than estimated, so nothing here is ever a guess.")
    _table(ws, 4,
           ["Date", "Wk", "Away", "Home", "Away Pts", "Home Pts", "Status", "Neutral",
            "Spread (home)", "Total", "Away ML", "Home ML", "Book"],
           [[(g.get("date") or "")[:10], g.get("week"), g.get("away"), g.get("home"),
             g.get("away_score"), g.get("home_score"),
             "Final" if g.get("completed") else (g.get("status") or "Upcoming"),
             "Y" if g.get("neutral") else "", (g.get("odds") or {}).get("spread_home"),
             (g.get("odds") or {}).get("total"), (g.get("odds") or {}).get("ml_away"),
             (g.get("odds") or {}).get("ml_home"), (g.get("odds") or {}).get("book")]
            for g in games],
           [11, 5, 9, 9, 9, 9, 12, 8, 13, 8, 9, 9, 14],
           [None, "0", None, None, "0", "0", None, None, "+0.0;-0.0", "0.0", "+0;-0", "+0;-0", None],
           name="GamesTable")

    # ------------------------------------------------------- Settings (record)
    ws = wb.create_sheet("Settings")
    _title(ws, "Settings Used For This Run",
           "READ ONLY. This records what produced the file. Editing a value here changes nothing — "
           "the model runs in the pipeline, not in Excel. To change the model, edit "
           "config/settings.json and re-run. (Cell B5 is the one exception: the Dashboard and Ledger "
           "reference it as the bankroll baseline.)")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 92
    flat = [
        ("Starting Bankroll", cfg.get("bankroll", {}).get("starting"), "Bankroll the ledger starts from."),
        ("Kelly Fraction", cfg.get("bankroll", {}).get("kelly_fraction"), "Fraction of full Kelly used for every stake."),
        ("Max Stake % of Bankroll", cfg.get("bankroll", {}).get("max_stake_pct"), "Hard cap on any single stake."),
        ("Size Stakes Off", cfg.get("bankroll", {}).get("size_off"), "'settled' sizes off settled results only, which avoids a stake depending on its own outcome."),
        ("Best Bet threshold", cfg.get("tiers", {}).get("best_bet"), "Edge required for the top tier."),
        ("Good threshold", cfg.get("tiers", {}).get("good"), "Edge required for GOOD."),
        ("Lean threshold", cfg.get("tiers", {}).get("lean"), "Edge required for LEAN."),
        ("Selection haircut", cfg.get("model", {}).get("selection_haircut"), "Subtracted from every edge before tiering, for the winner's curse."),
        ("Market Blend", cfg.get("model", {}).get("market_blend"), "Weight on the de-vigged market probability. 0.50 minimised out-of-sample calibration error in walk-forward simulation."),
        ("Margin Std Dev (pts)", cfg.get("model", {}).get("margin_sd"), "Spread of FBS scoring margin around the projection."),
        ("Total Std Dev (pts)", cfg.get("model", {}).get("total_sd"), "Spread of combined score around the projection."),
        ("Key Numbers", "on" if cfg.get("model", {}).get("use_key_numbers") else "off", "Discrete margin distribution bumped at 3, 7, 10, 14 rather than a smooth normal curve."),
        ("Max Model Probability", cfg.get("model", {}).get("max_model_prob"), "Cap applied at the staking step so one input cannot produce an extreme bet."),
        ("Max bets per game", cfg.get("filters", {}).get("max_bets_per_game"), "Correlation guard — a team's ML and its spread are close to the same bet."),
        ("Max plays per week", cfg.get("filters", {}).get("max_plays_per_week"), "Volume cap, best edges first."),
        ("Ridge lambda", cfg.get("ratings", {}).get("ridge_lambda"), "Rating shrinkage. Higher pulls thin-evidence teams harder toward average."),
        ("MOV cap", cfg.get("ratings", {}).get("mov_cap"), "Margin cap so blowouts do not dominate the ratings fit."),
        ("Recency half-life (games)", cfg.get("ratings", {}).get("recency_halflife_games"), "Measured in team-games; multiply by 7 for days."),
        ("Solved Home Field (pts)", meta.get("home_field_advantage"), "Estimated from this season's results, shrunk toward the configured prior."),
        ("League avg points", meta.get("league_avg_points"), "Per team per game, used as the baseline for the totals model."),
        ("Odds source", ", ".join(cfg.get("data", {}).get("odds_provider_priority", [])[:3]), "ESPN's public feed. Free, keyless, single-book — no line shopping and no multi-book consensus."),
    ]
    _table(ws, 4, ["Setting", "Value", "What it does"],
           [[k, v, d] for k, v, d in flat], [34, 16, 92])
    for i in range(len(flat)):
        ws.cell(row=5 + i, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws["B5"].font = Font(name="Arial", size=10, bold=True, color="0000FF")
    ws["B5"].fill = FILL_NOTE

    # --------------------------------------------------------- Model Health
    ws = wb.create_sheet("Model Health")
    _title(ws, "Model Health", "Whether the model deserves to be believed. Read this before the Board.")
    cal = summary.get("calibration", [])
    r = _table(ws, 4, ["Claimed probability", "Bets", "Predicted", "Actual", "Gap"],
               [[c["bucket"], c["n"], c["predicted"], c["actual"], c["actual"] - c["predicted"]]
                for c in cal],
               [22, 9, 12, 12, 12],
               [None, "0", pctf, pctf, "+0.0%;-0.0%"])
    r += 2
    for label, val, fmt in [
        ("Average CLV", summary.get("avg_clv"), "+0.0%;-0.0%"),
        ("Bets beating the close", summary.get("clv_positive_rate"), pctf),
        ("Brier score", meta.get("brier"), "0.0000"),
        ("Settled bets", summary.get("settled"), "0"),
    ]:
        ws.cell(row=r, column=1, value=label).font = F_BOLD
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = fmt
        c.font = F_NUM
        r += 1
    r += 1
    c = ws.cell(row=r, column=1, value=(
        "What would tell you this model is broken. A calibration gap that stays negative across every "
        "bucket past 100 bets means it is systematically overconfident — raise market_blend. Average CLV "
        "below zero means it is consistently taking worse numbers than the market closes at, which no win "
        "rate can rescue. Tier ROI running backwards — LEAN out-earning BEST BET — means the edge estimate "
        "is not ordering bets correctly. Any of the three is a reason to stop betting it and go fix it."))
    c.font = F_BODY
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.fill = FILL_NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 4, end_column=5)

    # ------------------------------------------------------------ Model History
    ws = wb.create_sheet("Model History")
    _title(ws, "Model History",
           "Every market the model has ever priced, graded — whether or not it was ever bet. The Bet "
           "Ledger only shows games the model disagreed with the market on, which is exactly where its "
           "own error runs largest; this is the much larger, much less selected sample. One row is kept "
           "per market per game (the side the model actually favoured), so a win rate here is a real "
           "accuracy figure, not two complementary sides cancelling out to 50%.")
    r = 4
    for label, val, fmt in [
        ("Total logged", history.get("total_logged"), "0"),
        ("Settled", history.get("settled"), "0"),
        ("Pending", history.get("pending"), "0"),
        ("Brier (all settled)", history.get("brier"), "0.0000"),
    ]:
        ws.cell(row=r, column=1, value=label).font = F_BOLD
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = fmt
        c.font = F_KPI
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="FULL-RECORD CALIBRATION").font = F_BOLD
    r += 1
    hcal = history.get("calibration", [])
    r = _table(ws, r, ["Claimed probability", "N", "Predicted", "Actual", "Gap"],
               [[c["bucket"], c["n"], c["predicted"], c["actual"], c["actual"] - c["predicted"]]
                for c in hcal],
               [22, 9, 12, 12, 12],
               [None, "0", pctf, pctf, "+0.0%;-0.0%"])
    r += 2

    ws.cell(row=r, column=1, value="ACCURACY BY TIER").font = F_BOLD
    r += 1
    by_tier = history.get("by_tier", {})
    tier_order = ["BEST BET", "GOOD", "LEAN", "PASS"]
    tier_rows = [[t, by_tier[t]["n"], by_tier[t]["win_rate"]] for t in tier_order if t in by_tier]
    r = _table(ws, r, ["Tier", "N", "Win Rate"], tier_rows, [12, 8, 11], [None, "0", pctf])
    r += 2

    ws.cell(row=r, column=1, value="ACCURACY BY MARKET").font = F_BOLD
    r += 1
    by_market = history.get("by_market", {})
    mkt_rows = [[m, by_market[m]["n"], by_market[m]["win_rate"], by_market[m].get("brier")]
                for m in ("ATS", "ML", "TOTAL") if m in by_market]
    r = _table(ws, r, ["Market", "N", "Win Rate", "Brier"], mkt_rows,
               [12, 8, 11, 10], [None, "0", pctf, "0.0000"])
    r += 2

    ws.cell(row=r, column=1, value="WEEK TREND").font = F_BOLD
    r += 1
    wk_rows = [[w.get("week"), w.get("n"), w.get("win_rate"), w.get("brier")]
               for w in history.get("week_trend", [])]
    r = _table(ws, r, ["Week", "Predictions Settled", "Win Rate", "Brier"], wk_rows,
               [10, 20, 11, 10], [None, "0", pctf, "0.0000"])
    r += 2
    c = ws.cell(row=r, column=1, value=(
        "Watch this trend rather than any single week — a lower Brier and a win rate holding near its "
        "claimed probability as the season goes on is the model actually working, not luck on a handful "
        "of early games."))
    c.font = F_BODY
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.fill = FILL_NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=5)

    wb.save(path)
    return path


def main() -> int:
    out = os.path.join(ROOT, "NCAAF_Edge_Model.xlsx")
    build(out)
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
